import requests
import json
import os
import re
import datetime
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def requests_retry_session(
    retries=3,
    backoff_factor=0.3,
    status_forcelist=(500, 502, 504),
    session=None,
):
    session = session or requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff_factor,
        status_forcelist=status_forcelist,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    return session


def save_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    f = open(path, 'w')
    json.dump(data, f)
    f.close()


def save_file(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    f = open(path, 'w')
    f.write(data)
    f.close()


def get_json_file(data_file, data_type=dict):
    try:
        return json.load(open(data_file, 'r'))
    except OSError:
        return data_type()


def get_file(data_file, data_type=dict):
    try:
        return open(data_file, 'r').read()
    except OSError:
        return data_type()


def make_workbook(sheet_name=None):
    workbook = Workbook()
    sheet = workbook.active
    if sheet_name:
        sheet.title = sheet_name
    return workbook, sheet


def save_workbook(workbook, filename):
    try:
        workbook.save(filename)
        workbook.close()
    except PermissionError:
        print('无法写入{}，请检查是否在Excel中打开了文件，如有需要请另存为并关闭'.format(filename))


def timestamp_to_excel_date(ts):
    return datetime.datetime.fromtimestamp(ts)


def make_chunks(src, size):
    return [src[i * size:(i + 1) * size] for i in range((len(src) + size - 1) // size)]


def match_single(data, regexp, default=None, data_type=str):
    result = re.search(regexp, data)
    return data_type(result.group(1)) if result else default


def match_count(data, regexp):
    result = re.findall(regexp, data)
    return len(result) if result else 0


def match_date(data, regexp):
    result = match_single(data, regexp)
    return datetime.datetime.strptime(result, '%Y-%m-%d %H:%M') if result else None


def set_columns_width(sheet, columns_width):
    for index, width in enumerate(columns_width):
        sheet.column_dimensions[get_column_letter(index + 1)].width = width
